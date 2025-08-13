# -*- coding: utf-8 -*-
import os
import sys
from flask import Flask, request, render_template, send_file
from collections import Counter
import matplotlib.pyplot as plt
import tempfile
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import logging

app = Flask(__name__)
# Configura el logger para que los errores se muestren en la consola de Vercel
logging.basicConfig(level=logging.INFO)

# --- Funciones de Procesamiento ---
def procesar_after_hours(ruta_archivo):
    """
    Procesa el archivo para generar una gráfica semanal y la inserta en una nueva hoja.
    Este es el script para 'After Hours'.
    Retorna la ruta del archivo procesado.
    """
    try:
        wb = openpyxl.load_workbook(ruta_archivo)
        hoja = wb["Sheet0"]
    except KeyError:
        return None, "ERROR: No se encontró la hoja 'Sheet0'."
    
    if 'Reporte' in wb.sheetnames:
        wb.remove(wb['Reporte'])
    hoja_reporte = wb.create_sheet('Reporte')

    # Búsqueda de datos y procesamiento
    timestamp_col, timestamp_row = None, None
    for row in hoja.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value == "TIMESTAMP":
                timestamp_col = get_column_letter(cell.column)
                timestamp_row = cell.row
                break
        if timestamp_col:
            break
    
    if not timestamp_col:
        return None, "ERROR: No se encontró la columna 'TIMESTAMP'."

    llamadas = []
    for row in hoja.iter_rows(min_row=timestamp_row + 1, min_col=hoja[timestamp_col + '1'].column, max_col=hoja[timestamp_col + '1'].column):
        cell = row[0]
        if cell.value:
            llamadas.append(cell.value)
    
    dias = []
    for item in llamadas:
        if isinstance(item, str):
            if len(item) >= 24:
                dias.append(f"{item[0:6]}, {item[-8:-6]}")
    
    # Agrupar datos por día de la semana
    dias_por_semana = {'Mon': [], 'Tue': [], 'Wed': [], 'Thu': [], 'Fri': [], 'Sat': [],'Sun': []}
    for item in dias:
        partes = item.split(',')
        if len(partes) > 0:
            dia_semana = partes[0].strip()
            if dia_semana in dias_por_semana:
                dias_por_semana[dia_semana].append(item)

    etiquetas_semana = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    valores_semana = [sum(Counter(dias_por_semana[dia]).values()) for dia in etiquetas_semana]
    
    # Creación del gráfico
    plt.figure(figsize=(10, 3))
    plt.bar(etiquetas_semana, valores_semana, color=['red', 'green', 'purple', 'blue', 'pink', 'brown', 'orange'])
    plt.title('Total Missed Calls - Week Overview (After Hours)')
    plt.xticks(rotation=45)
    plt.grid(axis='y', linestyle='--', alpha=0.6)
    plt.tight_layout()
    
    img_path = os.path.join(tempfile.gettempdir(), 'week_graph_after_hours.png')
    plt.savefig(img_path)
    plt.close()
    
    img = Image(img_path)
    hoja_reporte.add_image(img, 'B2')
    
    # Limpieza de archivo temporal
    os.remove(img_path)
    
    ruta_salida = os.path.join(tempfile.gettempdir(), f"procesado_after_hours_{os.path.basename(ruta_archivo)}")
    wb.save(ruta_salida)
    
    return ruta_salida, None

def procesar_caller_disconnected(ruta_archivo):
    """
    Procesa el archivo para generar una gráfica diaria para cada día con llamadas.
    Este es el script para 'Caller Disconnected'.
    Retorna la ruta del archivo procesado.
    """
    try:
        wb = openpyxl.load_workbook(ruta_archivo)
        hoja = wb["Sheet0"]
    except KeyError:
        return None, "ERROR: No se encontró la hoja 'Sheet0'."
    
    if 'Reporte' in wb.sheetnames:
        wb.remove(wb['Reporte'])
    hoja_reporte = wb.create_sheet('Reporte')
    
    # Búsqueda de datos y procesamiento
    timestamp_col, timestamp_row = None, None
    for row in hoja.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value == "TIMESTAMP":
                timestamp_col = get_column_letter(cell.column)
                timestamp_row = cell.row
                break
        if timestamp_col:
            break
    
    if not timestamp_col:
        return None, "ERROR: No se encontró la columna 'TIMESTAMP'."

    llamadas = []
    for row in hoja.iter_rows(min_row=timestamp_row + 1, min_col=hoja[timestamp_col + '1'].column, max_col=hoja[timestamp_col + '1'].column):
        cell = row[0]
        if cell.value:
            llamadas.append(cell.value)
    
    dias = []
    for item in llamadas:
        if isinstance(item, str):
            if len(item) >= 24:
                dias.append(f"{item[0:6]}, {item[-8:-6]}")
    
    dias_por_semana = {'Mon': [], 'Tue': [], 'Wed': [], 'Thu': [], 'Fri': [], 'Sat': [],'Sun': []}
    for item in dias:
        partes = item.split(',')
        if len(partes) > 0:
            dia_semana = partes[0].strip()
            if dia_semana in dias_por_semana:
                dias_por_semana[dia_semana].append(item)
    
    fila_actual = 2
    colores_por_dia = {
        'Sun': 'orange', 'Mon': 'red', 'Tue': 'green', 'Wed': 'purple',
        'Thu': 'blue', 'Fri': 'pink', 'Sat': 'brown'
    }

    for dia, lista_dia in dias_por_semana.items():
        if not lista_dia:
            continue
        
        conteo_dia = Counter(lista_dia)
        etiquetas = list(conteo_dia.keys())
        valores = list(conteo_dia.values())
        
        if not etiquetas:
            continue

        etiquetas_am_pm = []
        numero = ""
        for etiqueta in etiquetas:
            partes = etiqueta.split(',')
            if len(partes) > 2:
                numero = partes[1].strip()
                hora_24 = int(partes[2].strip())
                
                hora_12 = ""
                if hora_24 == 0:
                    hora_12 = '12am'
                elif 1 <= hora_24 < 12:
                    hora_12 = f'{hora_24}am'
                elif hora_24 == 12:
                    hora_12 = '12pm'
                else:
                    hora_12 = f'{hora_24 - 12}pm'
                etiquetas_am_pm.append(hora_12)
        
        plt.figure(figsize=(10, 3))
        plt.bar(etiquetas_am_pm, valores, color=colores_por_dia.get(dia, 'gray'))
        
        for i, valor in enumerate(valores):
            plt.text(i, valor + 0.2, str(valor), ha='center', va='bottom', fontsize=8)

        plt.grid(axis='y', linestyle='--', alpha=0.6)
        plt.ylim(0, max(valores) * 1.5)
        plt.subplots_adjust(bottom=0.20)
        
        plt.xlabel('Time')
        plt.ylabel('Call count')
        plt.title(f'{dia} {numero} - Total missed calls: {sum(valores)} (Caller Disconnected)')
        plt.xticks(rotation=90)
        plt.tight_layout()

        img_path = os.path.join(tempfile.gettempdir(), f'graph_{dia}_caller_disconnected.png')
        plt.savefig(img_path)
        plt.close()
        
        img = Image(img_path)
        hoja_reporte.add_image(img, f'B{fila_actual}')
        os.remove(img_path)
        
        fila_actual += 20
        
    ruta_salida = os.path.join(tempfile.gettempdir(), f"procesado_caller_disconnected_{os.path.basename(ruta_archivo)}")
    wb.save(ruta_salida)
    
    return ruta_salida, None

# --- Rutas de Flask ---
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return "No se encontró el archivo en la solicitud.", 400
    
    file = request.files['file']
    if file.filename == '':
        return "No se seleccionó ningún archivo.", 400
    
    if file:
        script_a_usar = request.form.get('script')
        
        # Guardar el archivo temporalmente
        temp_dir = tempfile.gettempdir()
        ruta_archivo = os.path.join(temp_dir, file.filename)
        file.save(ruta_archivo)
        
        ruta_salida = None
        error_msg = None

        if script_a_usar == 'after_hours':
            ruta_salida, error_msg = procesar_after_hours(ruta_archivo)
        elif script_a_usar == 'caller_disconnected':
            ruta_salida, error_msg = procesar_caller_disconnected(ruta_archivo)
        else:
            error_msg = "Script no válido."
        
        # Limpiar el archivo subido
        os.remove(ruta_archivo)
        
        if error_msg:
            return f"Error: {error_msg}", 500
        
        # Enviar el archivo procesado al usuario
        return send_file(
            ruta_salida,
            as_attachment=True,
            download_name=os.path.basename(ruta_salida)
        )

if __name__ == '__main__':
    # Esto es solo para pruebas locales. En Vercel, no se ejecuta esta parte.
    app.run(debug=True)
